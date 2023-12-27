from faker import Faker
from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active

fake = Faker(['en_US'])

# for i in range(1, 16):
#     for j in range(1, 4):
#         ws.cell(row=i, column=1).value = fake.name()
#         ws.cell(row=i, column=2).value = fake.email()
#         ws.cell(row=i, column=3).value = fake.cryptocurrency_name()
#         ws.cell(row=i, column=4).value = fake.job()
# wb.save("testData.xlsx")
b = fake.unique.first_name()
a = fake.word()
c = fake.street_name()
d = fake.sentence(nb_words=10)
e = fake.name()
f = fake.user_name()
g = fake.random_int(min=1, max=100)
h = fake.street_name()

print(h)

